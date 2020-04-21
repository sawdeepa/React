import openpyxl,datetime,copy,xlsxwriter,getpass,pysftp,paramiko,os,sys
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from operator import itemgetter

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from zipfile import ZipFile
import warnings
import time
import win32com.client as win32

import colorama
from colorama import Fore, Back, Style

colorama.init()
#os.system("mode con cols=150 lines=180")



chromeOptions = webdriver.ChromeOptions()
chromeOptions.add_argument("--log-level=3")
chromeOptions.add_experimental_option('useAutomationExtension', False)



chromeOptions1 = webdriver.ChromeOptions()
chromeOptions1.add_argument("--log-level=3")
chromeOptions1.add_argument("--start-maximized")
#chromeOptions1.add_argument("--disable-infobars")
chromeOptions1.add_experimental_option('useAutomationExtension', False)
path_read = r"C:\Users\deesaw\Desktop\ML\Machine Learning A-Z New\Part 2 - Regression\Section 5 - Multiple Linear Regression\Conversions_Validation_Package_Names_Consolidated.xlsx"
#path_read = r"D:\Conversions_Validation_Package_Names_Consolidated.xlsx"

wb = openpyxl.load_workbook(path_read,read_only=True)
#wb = openpyxl.load_workbook(path_read,read_only=True,guess_types=False)
sheet_pre = wb["PreLoad"]
sheet_post = wb["PostLoad"]
sheet_email = wb["Email_ID"]

max_column = 5

max_rows_pre = 79
max_rows_post = 71
dict_email_id = {}


total_rows_email = []
single_row = []
count=0
for row in sheet_email.iter_rows(max_col=2):#, max_row = max_rows_pre):
    count=count+1
    if row[0].value is None or str(row[0].value) == '':
        break
    for col in range(0,2,1):
        single_row.append(str(row[col].value).strip(" ").strip(",").strip("\n"))
    
    if single_row[0] not in dict_email_id:
        dict_email_id[single_row[0]] = single_row[1]
    single_row=[]


total_rows_pre = []
single_row = []
count=0
total_rows_pre.append([])
for row in sheet_pre.iter_rows(max_col=max_column):#, max_row = max_rows_pre):
    count=count+1
    if row[0].value is None or str(row[0].value) == '':
        break
    for col in range(0,max_column,1):
        single_row.append(row[col].value)
    total_rows_pre.append(single_row)
    single_row=[]


total_rows_post = []
single_row = []
count=0
total_rows_post.append([])
for row in sheet_post.iter_rows(max_col=max_column):#, max_row = max_rows_post):
    count=count+1
    if row[0].value is None or str(row[0].value) == '':
        break
    for col in range(0,max_column,1):
        single_row.append(row[col].value)
    total_rows_post.append(single_row)
    single_row=[]

for item in total_rows_pre[1:]:
    for i in range(len(item)):
        if item[i] is not None:
            item[i] = str(item[i]).replace("\n"," ").strip(' ')
        if item[i] == '':
            item[i] = None

for item in total_rows_post[1:]:
    for i in range(len(item)):
        if item[i] is not None:
            item[i] = str(item[i]).replace("\n"," ").strip(' ')
        if item[i] == '':
            item[i] = None


dict_pre_obj = {}
dict_post_obj = {}
dict_etl = {}


for item in total_rows_pre[2:]:
    
    if item[1] is not None and item[1] in dict_pre_obj:
        l = dict_pre_obj[item[1]]
        l[0] = l[0] + 1
        l.append([item[0],item[2],item[3],item[4]])
    
    if item[1] is not None and item[1] not in dict_pre_obj:
        c = []
        dict_pre_obj[item[1]] = c
        c.append(1)
        c.append([item[0],item[2],item[3],item[4]])
        c=[]

    if item[0] is not None and item[0] in dict_etl:
        l = dict_etl[item[0]]
        l[0] = l[0] + 1
        l.append([item[1],item[2],item[3],item[4],"PRE"])

    if item[0] is not None and item[0] not in dict_etl:
        c = []
        dict_etl[item[0]] = c
        c.append(1)
        c.append([item[1],item[2],item[3],item[4],"PRE"])
        c=[]


for item in total_rows_post[2:]:
    
    if item[1] is not None and item[1] in dict_post_obj:
        l = dict_post_obj[item[1]]
        l[0] = l[0] + 1
        l.append([item[0],item[2],item[3],item[4]])
    
    if item[1] is not None and item[1] not in dict_post_obj:
        c = []
        dict_post_obj[item[1]] = c
        c.append(1)
        c.append([item[0],item[2],item[3],item[4]])
        c=[]

    if item[0] is not None and item[0] in dict_etl:
        l = dict_etl[item[0]]
        l[0] = l[0] + 1
        l.append([item[1],item[2],item[3],item[4],"POST"])

    if item[0] is not None and item[0] not in dict_etl:
        c = []
        dict_etl[item[0]] = c
        c.append(1)
        c.append([item[1],item[2],item[3],item[4],"POST"])
        c=[]


if sys.platform.lower() == "win32":
    os.system('color')

# Group of Different functions for different styles
class style():
    BLACK = lambda x: '\033[30m' + str(x)
    RED = lambda x: '\033[31m' + str(x)
    GREEN = lambda x: '\033[32m' + str(x)
    YELLOW = lambda x: '\033[33m' + str(x)
    BLUE = lambda x: '\033[34m' + str(x)
    MAGENTA = lambda x: '\033[35m' + str(x)
    CYAN = lambda x: '\033[36m' + str(x)
    WHITE = lambda x: '\033[37m' + str(x)
    UNDERLINE = lambda x: '\033[4m' + str(x)
    RESET = lambda x: '\033[0m' + str(x)

CEND      = '\33[0m'
CBOLD     = '\33[1m'
CITALIC   = '\33[3m'
CURL      = '\33[4m'
CBLINK    = '\33[5m'
CBLINK2   = '\33[6m'
CSELECTED = '\33[7m'

CBLACK  = '\33[30m'
CRED    = '\33[31m'
CGREEN  = '\33[32m'
CYELLOW = '\33[33m'
CBLUE   = '\33[34m'
CVIOLET = '\33[35m'
CBEIGE  = '\33[36m'
CWHITE  = '\33[37m'

CBLACKBG  = '\33[40m'
CREDBG    = '\33[41m'
CGREENBG  = '\33[42m'
CYELLOWBG = '\33[43m'
CBLUEBG   = '\33[44m'
CVIOLETBG = '\33[45m'
CBEIGEBG  = '\33[46m'
CWHITEBG  = '\33[47m'

CGREY    = '\33[90m'
CRED2    = '\33[91m'
CGREEN2  = '\33[92m'
CYELLOW2 = '\33[93m'
CBLUE2   = '\33[94m'
CVIOLET2 = '\33[95m'
CBEIGE2  = '\33[96m'
CWHITE2  = '\33[97m'

CGREYBG    = '\33[100m'
CREDBG2    = '\33[101m'
CGREENBG2  = '\33[102m'
CYELLOWBG2 = '\33[103m'
CBLUEBG2   = '\33[104m'
CVIOLETBG2 = '\33[105m'
CBEIGEBG2  = '\33[106m'
CWHITEBG2  = '\33[107m'

#print(os.getcwd())

user_name = None
passwd = None
hostname_qual = 'gplvapdsq1'
hostname_prod= 'gplvapdsp1'
sftp = None


"""with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            cnopts = pysftp.CnOpts()
            cnopts.hostkeys = None
            sftp = pysftp.Connection(hostname_qual, username=user_name, password=passwd,cnopts=cnopts)"""

etl_id = None
base_path = "/sapefsprod/CONV/General/D2/"
path_local = os.getcwd()
ele_type = None

phase_path = "7 - D2"
team_path = "Data"
file_path = None
driver1 = None


def winscp_login():
    global sftp
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            cnopts = pysftp.CnOpts()
            cnopts.hostkeys = None
            sftp = pysftp.Connection(hostname_prod, username=user_name, password=passwd,cnopts=cnopts)
    except:
        print("\n..  " + Fore.WHITE + Back.RED +  "Login Failed Due to incorrect User Name/Password. Please try again!" + Style.RESET_ALL +"\n")
        return(404)

aws_user = os.environ['USERNAME']

deloitte_email = None
if aws_user in dict_email_id:
    deloitte_email = dict_email_id[aws_user]

stext = None
subject = None

def Send_Email(base_path1,zip_name,adur,recon_path):
    global stext,subject
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = deloitte_email

    mail.Subject = subject

    mail.HTMLBody = """<p>Hi All,</p><p class=MsoNormal>Actual Time Duration : <b>""" + adur + """</b></p>

    <p class=MsoNormal><b>SharePoint Path:</b></p>

    <p class=MsoNormal>""" +stext + """</p>

    <ul style='margin-top:0in' type=disc>
    <li class=MsoListParagraph style='margin-left:0in;mso-list:l1 level1 lfo3;
    text-autospace:none'><b><span style='font-size:10.0pt;font-family:"Segoe UI",sans-serif;
    mso-fareast-font-family:"Times New Roman";color:black'>""" + zip_name +"""</span></b><b><span
    style='mso-fareast-font-family:"Times New Roman"'><o:p></o:p></span></b></li>
    </ul>
    <p class="MsoNormal" style="text-autospace:none"><b>WinSCP directory:</b></p>
    <p class="MsoNormal" style="text-autospace:none">""" + base_path1 + """ <o:p></o:p></p>"""
    
    if recon_path is not None:
        mail.Attachments.Add(recon_path)
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            mail.Send()
            recon_path = None
            stext = None
            subject = None
            print("..  " + Fore.BLACK + Back.GREEN + "Email sent Successfully to : " + deloitte_email + Style.RESET_ALL + "\n" )
    except:
        recon_path = None
        stext = None
        subject = None
        print("..  " + Fore.WHITE + Back.RED + "Error occured in Sending Email!" +Style.RESET_ALL +"\n")
        
def txt_to_excel(o):
    path_write = o.split("\\")[-1].strip("txt").strip("TXT")+"xlsx"
    total_rows = []
    total_rows=[]
    count=0
    with open(o,encoding="utf8") as f:
        for line in f:
            count  = count + 1
            h=[]
            h.append(line)
            k=h[0].strip('\n').split("\t")
            total_rows.append(k)
		
    workbook = xlsxwriter.Workbook(path_write)
    worksheet = workbook.add_worksheet()
    for row in range(0,len(total_rows),1):
        #print(row)
        for col in range(0,len(total_rows[row]),1):
            worksheet.write(row,col,total_rows[row][col])
    try:
        workbook.close()
        return(path_write)

    except:
        print("..  " + Fore.WHITE + Back.RED + "Error occurred in converting Profiling text file to Excel.Please try again!" +Style.RESET_ALL +"\n")
        return(None)
    
    
def upload_and_create():
    if sftp is None:
        while True:
            global user_name,passwd
            print("..  Enter your Winscp User Name OR Enter 0 to go back.")
            user_name = input(">>  ")
            if user_name == '0':
                return
            passwd = getpass.getpass(prompt="..  Enter your Winscp Password:\n>>  ")
            if winscp_login() != 404:
                print("..  " + Fore.BLACK + Back.GREEN + "Login Successful\n" + Style.RESET_ALL)
                break
    while True:
        global file_path,subject,stext
        print("..  Enter ETL ID only OR RICEF ID followed by Post/Pre: ")
        print("..  Enter 0 to go back\n")
        obj = input(">>  ")
        if obj == '0':
            return
        l = obj.split(' ')
        if len(l) < 1:
            print("..  " + Fore.WHITE + Back.RED + "Wrong Input! Enter again!" +Style.RESET_ALL +"\n")
            continue
        if len(l) == 1:
            if l[0] in (dict_post_obj) or l[0] in (dict_pre_obj):
                print("..  " + Fore.WHITE + Back.RED + "Wrong Input! Please Enter POST/PRE after RICEF ID.For ex: ITR_C0016 Post" +Style.RESET_ALL +"\n")
            elif l[0] not in dict_etl:
                print("..  " + Fore.WHITE + Back.RED + "Wrong ETL ID! Enter again!" +Style.RESET_ALL +"\n")
                continue
            elif l[0] in dict_etl:
                """ABCD"""
                recon_path = None
                temp_list = dict_etl[l[0]]
                one_time = None
                for i in range(0,temp_list[0]):
                    
                    break_label = None
                    var1 = temp_list[i+1]
                    part_obj = var1[3]
                    recon_path = None

                    if temp_list[0] > 1:

                        print("\n..  Validation Package for the following zip file is going to be created now: \n    "+ Back.BLUE+ var1[1] + Style.RESET_ALL + "\n")
                        print("..  Enter 0 to skip and proceed with other Validation packages for the same ETL ID or \n    press any other key to continue with the above pacakge. "+Back.RED +"Press 1 to Quit."+Style.RESET_ALL+"\n")

                        s1 = input(">>  ")

                        if s1 == '0':
                            continue
                        if s1 == '1':
                            return
                    

                    if var1[2] is not None:
                        base_path1 = base_path + var1[2]

                    if var1[2] is None:
                        
                        print("..  " + Fore.WHITE + Back.RED + "Winscp Path is not there in the tracker for : \n    "+ Back.BLUE+ var1[1] + Style.RESET_ALL + "\n")
                        print("..  " + Fore.WHITE + Back.RED + "Enter Winscp path manually for creating the above zip package or 0 to continue." + Style.RESET_ALL + "\n")
                        while True:

                            base_path1 = input(">>  ")
                            if base_path1 == '0':
                                break_label = 0
                                break
                            if 'SAPEFS' not in base_path1.upper():
                                base_path1 = base_path + base_path1
                            
                            if sftp.isdir(base_path1):
                                break
                            
                            print("..  " + Fore.WHITE + Back.RED + "Winscp Path entered is wrong "+ Style.RESET_ALL + "\n")
                            print("..  " + Fore.WHITE + Back.RED + "Enter Winscp path manually for creating the above zip package or 0 to continue." + Style.RESET_ALL + "\n")

                            

                    if break_label == 0:
                        continue      

                    if not sftp.isdir(base_path1):
                        print("..  " + Fore.WHITE + Back.RED + "Winscp Path in tracker for : \n    "+ Back.BLUE+  var1[1] +Back.RED+ " Zip file is wrong."+ Style.RESET_ALL + "\n")
                        print("..  " + Fore.WHITE + Back.RED + "Enter Winscp path manually for creating the above zip package or 0 to continue." + Style.RESET_ALL + "\n")

                        while True:
                            
                            base_path1 = input(">>  ")
                            if base_path1 == '0':
                                break_label = 0
                                break
                            if 'SAPEFS' not in base_path1.upper():
                                base_path1 = base_path + base_path1
                            
                            if sftp.isdir(base_path1):
                                break
                            
                            print("..  " + Fore.WHITE + Back.RED + "Winscp Path entered is wrong "+ Style.RESET_ALL + "\n")
                            print("..  " + Fore.WHITE + Back.RED + "Enter Winscp path manually for creating the above zip package or 0 to continue." + Style.RESET_ALL + "\n")   

                    if break_label == 0:
                        continue

                    if sftp.isdir(base_path1):
                        with sftp.cd(base_path1):
                            list_dir = sftp.listdir()
                            if list_dir == []:
                                 print("..  " + Fore.WHITE + Back.RED +" Winscp Folder is empty for "+Back.BLUE+ var1[1] + "\n" + Style.RESET_ALL)
                                 continue

                            temp = []
                            break_label = None
                            
                            for li in list_dir:
                                
                                if sftp.isfile(li):
                                    if li.split("_")[0] == var1[1].split("_")[0]:
                                        sftp.get(li,preserve_mtime=True)
                                        o = os.getcwd()+"\\"+li
                                        if os.path.isfile(o):
                                            if "RECON" in li.upper() and ".XLSX" in li.upper():
                                                recon_path = o
                                            if "SUCCESSPROFILE" in li.upper() and ".TXT" in li.upper():
                                                o1 = txt_to_excel(o)
                                                if o1 is not None:
                                                    o = os.getcwd()+"\\"+ o1
                                                    li = o1
                                                    try:
                                                        sftp.put(o)
                                                        sftp.chmod(o1,777)
                                                    except:
                                                        pass
                                                
                                                if o1 is None:
                                                    break_label = 0
                                                    break
                                            temp.append(li)
                                            
                                        else:
                                            print("..  " + Fore.WHITE + Back.RED +str(li) + " not extracted from Winscp. Please try again!" + Style.RESET_ALL)
                                            break_label = 0
                                            break
                                        
                        if break_label == 0:
                            continue

                        if temp == []:
                            print("..  " + Fore.WHITE + Back.RED +" Winscp Folder is empty for "+Back.BLUE+ var1[1] + "\n" + Style.RESET_ALL)

                        print("..  These files will be converted into ZIP:\n")
                        for li in temp:
                            p = os.path.basename(li)
                            print("    * " + Fore.BLACK + Back.WHITE + p + Style.RESET_ALL)

                        time1 = datetime.datetime.now().strftime('%d%m%y%H%M%S')
                        zip_name = var1[1]
                        zip_name = "_".join(zip_name.split("_")[:-1]) + "_" + time1 + ".zip"

                        with ZipFile(zip_name,"w") as new_zip:
                            for t in temp:
                                new_zip.write(t)

                        print("\n\n..  Validation Pacakge Zip has been created with the following name:\n..  " + Fore.BLACK + Back.WHITE + zip_name + Style.RESET_ALL)

                        if part_obj is None:
                            print("..  " + Fore.WHITE + Back.RED +" Sharepoint Folder is not available. Please use Option 2 in main menu to upload the zip file" + Style.RESET_ALL)
                            continue                      
    
                        if part_obj.upper() not in ('AC','CTM','DAT','NTO','OTC','PTP','RTM','SCM','WM','MAT'):
                            print("..  " + Fore.WHITE + Back.RED + part_obj + " Sharepoint Folder is incorrect. Please use Option 2 in main menu to upload it" + Style.RESET_ALL)
                            continue

                        print("\n..  It will be uploaded to " + part_obj + " Folder in Sharepoint\n")
                        print("..  If Zip File name is incorrect or Sharepoint Folder is incorrect:\n")
                        print("..  Enter 0 and rename it and upload it via option 2 in main_menu")
                        print("..  Else press any other key to proceed with the upload\n")
                        choice = input(">>  ")

                        if choice == '0':
                            continue
                        
                        else:
                            if deloitte_email is not None:
                                while True:
                                    print("\n..  Enter Actual Duration for task in hours\n")
                                    aduration = input(">>  ")
                                    if aduration is None:
                                        print("..  " + Fore.WHITE + Back.RED +" Please enter Actual Task Duration!" + Style.RESET_ALL)
                                    if aduration is not None:
                                        if 'hour'.upper() not in aduration.upper() or 'hrs'.upper not in aduration.upper():
                                            if aduration.split(' ')[0] == '1':
                                                aduration = aduration + ' hour'
                                                
                                            elif aduration.split(' ')[0] != '1':
                                                aduration = aduration + ' hours'
                                        break

                                #global subject,stext
                                subject = "Details for Task - " + var1[0].upper()
                                stext = "Data team docs -> D2 -> D2 - Load and validation summary ->" + part_obj
                                Send_Email(base_path1,zip_name,aduration,recon_path)

                                    
                                
                            #global file_path
                            file_path = os.getcwd()+"\\"+zip_name

                            if one_time is None:
                                print("..  While the ZIP File is being uploaded, Please enter the online Recon\n")
                                time.sleep(1)
                                upload_recon(var1[0])
                                one_time = 0

                            if part_obj == 'AC':
                                upload_only('4')
                                
                            if part_obj == 'ITR':
                                upload_only('4')

                            if part_obj == 'CTM':
                                upload_only('5')

                            if part_obj == 'DAT':
                                upload_only('6')                                
                            
                            if part_obj == 'NTO':
                                upload_only('7')                        
                                
                            if part_obj == 'OTC':
                                upload_only('8')
                                
                            if part_obj == 'PTP':
                                upload_only('9')
                                
                            if part_obj == 'RTM':
                                upload_only('10')

                            if part_obj == 'SCM':
                                upload_only('11')

                            if part_obj == 'WM':
                                upload_only('12')

                            if part_obj  == 'MAT':
                                upload_only('1')
                            
                    
                
        else:
            if str(l[-1]).upper() not in ('POST','PRE'):
               print("..  " + Fore.WHITE + Back.RED + "Wrong Input! Please Enter POST/PRE after RICEF ID.For ex: ITR_C0016 Post" +Style.RESET_ALL +"\n")
               continue

            if (str(l[-1]).upper() == "PRE" and str(l[0]).upper() not in dict_pre_obj) or (str(l[-1]).upper() == "POST" and str(l[0]).upper() not in dict_post_obj):
                print("..  " + Fore.WHITE + Back.RED + "Wrong RICEF ID Or RICEF ID Post/Pre Combination is not Valid! Please try again" +Style.RESET_ALL +"\n")
                continue


            if str(l[-1]).upper() == "PRE":
                recon_path = None
                l[0]=l[0].upper()
                if dict_pre_obj[l[0]][0] > 1:
                    print("..  " + Fore.WHITE + Back.RED + "RICEF ID has following Multiple Tasks! Please enter ETL ID instead" +Style.RESET_ALL +"\n")
                    t0 = dict_pre_obj[l[0]]
                    print("    * " + Fore.WHITE + Back.BLUE +  "ETL TASK ID"+ Style.RESET_ALL+ "  -   "+Fore.WHITE + Back.BLUE+"Validation Package Name \n" + Style.RESET_ALL)
                    for i in range(1,len(t0),1):
                        print("    * " + Fore.BLACK + Back.WHITE + t0[i][0] + " - " + t0[i][1] + Style.RESET_ALL)
                    print("\n")
                        
                    continue
                part_obj = l[0].split("_")[0]
                
                if dict_pre_obj[l[0]][1][2] is not None:
                    base_path1 = base_path + dict_pre_obj[l[0]][1][2]
                else:
                    base_path1 = base_path + 'Validation_Report/' + part_obj + "/" + l[0] + "/" + "Preload_Report"
                    
                if not sftp.isdir(base_path1):
                    while True:
                        print("..  " + Fore.WHITE + Back.RED + "Winscp Path doesn't exist. Enter Winscp path manually or 0 to quit."+ Style.RESET_ALL + "\n")
                        base_path1 = input(">>  ")
                        if base_path1 == '0':
                            return
                        if 'SAPEFS' not in base_path1.upper():
                            base_path1 = base_path + base_path1
                            
                        if sftp.isdir(base_path1):
                            break
                    
                        
                if sftp.isdir(base_path1):
                    with sftp.cd(base_path1):
                        list_dir = sftp.listdir()
                        if list_dir == []:
                            print("..  " + Fore.WHITE + Back.RED +" Folder is empty, Please try again!" + Style.RESET_ALL)
                            return

                        for li in list_dir:
                            if sftp.isfile(li):
                                if '.' in li:
                                    t1 = li.split('.')
                                    t2 = t1[0]
                                    if '_' in t2:
                                        t3 = "_".join(t2.split("_")[:-1])+'.'+t1[1]
                                    t2.split
                                    #t2 = "_".join(zip_name.split("_")[:-1])

                        sftp.get_d(base_path1,path_local,preserve_mtime=True)
                        temp = []
                        for li in list_dir:
                            if sftp.isfile(li):
                                o = os.getcwd()+"\\"+li
                                if os.path.isfile(o):
                                    if "RECON" in li.upper() and ".XLSX" in li.upper():
                                        recon_path = o
                                    if "SUCCESSPROFILE" in li.upper() and ".TXT" in li.upper():
                                        o1 = txt_to_excel(o)
                                        if o1 is not None:
                                            o = os.getcwd()+"\\"+ o1
                                            li = o1
                                            try:
                                                sftp.put(o)
                                                sftp.chmod(o1,777)
                                            except:
                                                pass
                                            
                                        if o1 is None:
                                            return
                                    temp.append(li)
                                else:
                                    print("..  " + Fore.WHITE + Back.RED +str(li) + " not extracted from Winscp. Please try again!" + Style.RESET_ALL)
                                    return

                        if temp == []:
                            print("..  " + Fore.WHITE + Back.RED + " Folder is empty, Please try again!" + Style.RESET_ALL)
                            return

                        print("..  These files will be converted into ZIP:\n")
                        for li in temp:
                            p = os.path.basename(li)
                            print("    * " + Fore.BLACK + Back.WHITE + p + Style.RESET_ALL)
                            
                        time1 = datetime.datetime.now().strftime('%d%m%y%H%M%S')
                        zip_name = dict_pre_obj[l[0]][1][1]
                        zip_name = "_".join(zip_name.split("_")[:-1]) + "_" + time1 + ".zip"
                        #zip_name = dict_pre_obj[l[0]][1][1] + "_" + time + ".zip"
                        with ZipFile(zip_name,"w") as new_zip:
                            for t in temp:
                                #print(t)
                                new_zip.write(t)

                        part_obj = dict_pre_obj[l[0]][1][3]

                        print("\n\n..  Validation Pacakge Zip has been created with the following name:\n..  " + Fore.BLACK + Back.WHITE + zip_name + Style.RESET_ALL)

                        if part_obj is None:
                            print("..  " + Fore.WHITE + Back.RED +" Sharepoint Folder is not available. Please use Option 2 in main menu to upload the zip file" + Style.RESET_ALL)
                            return

                        if part_obj.upper() not in ('AC','CTM','DAT','NTO','OTC','PTP','RTM','SCM','WM','MAT'):
                            print("..  " + Fore.WHITE + Back.RED + part_obj + " Sharepoint Folder is incorrect. Please use Option 2 in main menu to upload it" + Style.RESET_ALL)
                            return

                        print("\n..  It will be uploaded to " + part_obj + " Folder in Sharepoint\n")
                        print("..  If Zip File name is incorrect or Sharepoint Folder is incorrect:\n")
                        print("..  Enter 0 and rename it and upload it via option 2 in main_menu")
                        print("..  Else press any other key to proceed with the upload\n")
                        choice = input(">>  ")
                        
                        if choice == '0':
                            return
                        else:
                            if deloitte_email is not None:
                                while True:
                                    print("\n..  Enter Actual Duration for task in hours\n")
                                    aduration = input(">>  ")
                                    if aduration is None:
                                        print("..  " + Fore.WHITE + Back.RED +" Please enter Actual Task Duration!" + Style.RESET_ALL)

                                    if aduration is not None:
                                        if 'hour'.upper() not in aduration.upper() or 'hrs'.upper not in aduration.upper():
                                            if aduration.split(' ')[0] == '1':
                                                aduration = aduration + ' hour'
                                                
                                            elif aduration.split(' ')[0] != '1':
                                                aduration = aduration + ' hours'
                                        break

                                #global subject,stext
                                subject = "Details for Task - " + l[0].upper()
                                stext = "Data team docs -> D2 -> D2 - Load and validation summary ->" + part_obj
                                #recon_path = os.getcwd() + "\\" + "5259_Pre_SourceList_ECC_SuccessProfile_290619094549.xlsx"
                                Send_Email(base_path1,zip_name,aduration,recon_path)

                                    
                                
                            #global file_path
                            file_path = os.getcwd()+"\\"+zip_name
                            
                            print("..  While the ZIP File is being uploaded, Please enter the online Recon\n")
                            time.sleep(1)
                            upload_recon(l[0])
                            
                            if part_obj == 'AC':
                                upload_only('4')
                                
                            if part_obj == 'ITR':
                                upload_only('4')

                            if part_obj == 'CTM':
                                upload_only('5')

                            if part_obj == 'DAT':
                                upload_only('6')                                
                            
                            if part_obj == 'NTO':
                                upload_only('7')                        
                                
                            if part_obj == 'OTC':
                                upload_only('8')
                                
                            if part_obj == 'PTP':
                                upload_only('9')
                                
                            if part_obj == 'RTM':
                                upload_only('10')

                            if part_obj == 'SCM':
                                upload_only('11')

                            if part_obj == 'WM':
                                upload_only('12')

                            if part_obj  == 'MAT':
                                upload_only('1')


            if str(l[-1]).upper() == "POST":
                recon_path = None
                l[0]=l[0].upper()
                if dict_post_obj[l[0]][0] > 1:
                    print("..  " + Fore.WHITE + Back.RED + "RICEF ID has following Multiple Tasks! Please enter ETL ID instead" +Style.RESET_ALL +"\n")
                    t0 = dict_post_obj[l[0]]
                    print("    * " + Fore.WHITE + Back.BLUE +  "ETL TASK ID"+ Style.RESET_ALL+ "  -   "+Fore.WHITE + Back.BLUE+"Validation Package Name \n" + Style.RESET_ALL)

                    for i in range(1,len(t0),1):
                        print("    * " + Fore.BLACK + Back.WHITE + t0[i][0] + " - " + t0[i][1] + Style.RESET_ALL)
                        
                    print("\n")
                        
                    continue
                part_obj = l[0].split("_")[0]
                
                if dict_post_obj[l[0]][1][2] is not None:
                    base_path1 = base_path + dict_post_obj[l[0]][1][2]
                else:
                    base_path1 = base_path + 'Validation_Report/' + part_obj + "/" + l[0] + "/" + "Postload_Report"
                    
                if not sftp.isdir(base_path1):
                    while True:
                        print("..  " + Fore.WHITE + Back.RED + "Winscp Path doesn't exist. Enter Winscp path manually or 0 to quit."+ Style.RESET_ALL + "\n")
                        base_path1 = input(">>  ")
                        if base_path1 == '0':
                            return
                        if 'SAPEFS' not in base_path1.upper():
                            base_path1 = base_path + base_path1
                            
                        if sftp.isdir(base_path1):
                            break
                    
                        
                if sftp.isdir(base_path1):
                    with sftp.cd(base_path1):
                        list_dir = sftp.listdir()
                        if list_dir == []:
                            print("..  " + Fore.WHITE + Back.RED +" Folder is empty, Please try again!" + Style.RESET_ALL)
                            return

                        for li in list_dir:
                            if sftp.isfile(li):
                                if '.' in li:
                                    t1 = li.split('.')
                                    t2 = t1[0]
                                    if '_' in t2:
                                        t3 = "_".join(t2.split("_")[:-1])+'.'+t1[1]
                                    t2.split
                                    #t2 = "_".join(zip_name.split("_")[:-1])
                        sftp.get_d(base_path1,path_local,preserve_mtime=True)
                        temp = []
                        for li in list_dir:
                            if sftp.isfile(li):
                                o = os.getcwd()+"\\"+li
                                if os.path.isfile(o):
                                    if "RECON".upper() in li.upper():
                                        recon_path = o
                                    temp.append(li)
                                else:
                                    print("..  " + Fore.WHITE + Back.RED +str(li) + " not extracted from Winscp. Please try again!" + Style.RESET_ALL)
                                    return

                        if temp == []:
                            print("..  " + Fore.WHITE + Back.RED +" Folder is empty, Please try again!" + Style.RESET_ALL)
                            return

                        print("..  These files will be converted into ZIP:\n")
                        for li in list_dir:
                            if sftp.isfile(li):
                                print("    * " + Fore.BLACK + Back.WHITE + li + Style.RESET_ALL)
                            
                        time1 = datetime.datetime.now().strftime('%d%m%y%H%M%S')
                        zip_name = dict_post_obj[l[0]][1][1]
                        zip_name = "_".join(zip_name.split("_")[:-1]) + "_" + time1 + ".zip"
                        #zip_name = dict_post_obj[l[0]][1][1] + "_" + time + ".zip"
                        with ZipFile(zip_name,"w") as new_zip:
                            for t in temp:
                                #print(t)
                                new_zip.write(t)

                        part_obj = dict_post_obj[l[0]][1][3]

                        print("\n\n..  Validation Pacakge Zip has been created with the following name:\n..  " + Fore.BLACK + Back.WHITE + zip_name + Style.RESET_ALL)

                        if part_obj is None:
                            print("..  " + Fore.WHITE + Back.RED +" Sharepoint Folder is not available. Please use Option 2 in main menu to upload it" + Style.RESET_ALL)
                            return

                        if part_obj.upper() not in ('AC','CTM','DAT','NTO','OTC','PTP','RTM','SCM','WM','MAT'):
                            print("..  " + Fore.WHITE + Back.RED + part_obj + " Sharepoint Folder is incorrect. Please use Option 2 in main menu to upload it" + Style.RESET_ALL)
                            return

                        print("\n..  It will be uploaded to " + part_obj + " Folder in Sharepoint\n")
                        print("..  If Zip File name is incorrect or Sharepoint Folder is incorrect:\n")
                        print("..  Enter 0 and rename it and upload it via option 2 in main_menu")
                        print("..  Else press any other key to proceed with the upload\n")
                        choice = input(">>  ")
                        
                        if choice == '0':
                            return
                        else:
                            if deloitte_email is not None:
                                while True:
                                    print("\n..  Enter Actual Duration for task in hours\n")
                                    aduration = input(">>  ")
                                    if aduration is None:
                                        print("..  " + Fore.WHITE + Back.RED +" Please enter Actual Task Duration!" + Style.RESET_ALL)

                                    if aduration is not None:
                                        if 'hour'.upper() not in aduration.upper() or 'hrs'.upper not in aduration.upper():
                                            if aduration.split(' ')[0] == '1':
                                                aduration = aduration + ' hour'
                                                
                                            elif aduration.split(' ')[0] != '1':
                                                aduration = aduration + ' hours'
                                                
                                        break

                                #global subject,stext
                                subject = "Details for Task - " + l[0].upper()
                                stext = "Data team docs -> D2 -> D2 - Load and validation summary ->" + part_obj
                                #recon_path = os.getcwd() + "\\" + "5259_Pre_SourceList_ECC_SuccessProfile_290619094549.xlsx"
                                Send_Email(base_path1,zip_name,aduration,recon_path)

                                    
                                
                            #global file_path
                            file_path = os.getcwd()+"\\"+zip_name
                            
                            print("..  While the ZIP File is being uploaded, Please enter the online Recon\n")
                            time.sleep(2)
                            upload_recon(l[0])
                            
                            if part_obj == 'AC':
                                upload_only('4')
                                
                            if part_obj == 'ITR':
                                upload_only('4')

                            if part_obj == 'CTM':
                                upload_only('5')

                            if part_obj == 'DAT':
                                upload_only('6')                                
                            
                            if part_obj == 'NTO':
                                upload_only('7')                        
                                
                            if part_obj == 'OTC':
                                upload_only('8')
                                
                            if part_obj == 'PTP':
                                upload_only('9')
                                
                            if part_obj == 'RTM':
                                upload_only('10')

                            if part_obj == 'SCM':
                                upload_only('11')

                            if part_obj == 'WM':
                                upload_only('12')

                            if part_obj  == 'MAT':
                                upload_only('1')
                                
def upload_only(choice):

    if choice == 0:
        print("..  Enter name of Zip File you want to Upload To Sharepoint.")
        print("..  Enter full path if file not in current Directory else you can enter only File Name.")
        print("..  Enter 0 if you don't want to upload a Zip File\n")
        global file_path
        file_path = input(">>  ")
        #print("\n")

        if file_path == '0':
            return

        if "\\" in file_path:
            if not os.path.isfile(file_path) or os.stat(file_path).st_size == 0:
                print("..  The file is empty/doesn't exist.Please try again!\n")
                upload_only(0)   

        elif "\\" not in file_path:
            file_path = os.getcwd()+"\\"+file_path
            if not os.path.isfile(file_path) or os.stat(file_path).st_size == 0:
                print("..  The file is empty/doesn't exist.Please try again!\n")
                upload_only(0)
                    
        while True:
                print("\n..  Enter 1 to Upload to Material Folder in Sharepoint")
                print("..  Enter 2 for Vendor")
                print("..  Enter 3 for Customer")
                print("..  Enter 4 for AC")
                print("..  Enter 5 for CTM")
                print("..  Enter 6 for DAT")
                print("..  Enter 7 for NTO")
                print("..  Enter 8 for OTC")
                print("..  Enter 9 for PTP")
                print("..  Enter 10 for RTR")
                print("..  Enter 11 for SCM")
                print("..  Enter 12 for WM")
                print("..  Enter 0 to Exit\n")
                choice = input(">>  ")

                if choice == '0':
                    return
                
                if choice not in ('1','2','3','4','5','6','7','8','9','10','11','12'):
                    print("..  Invalid Entry! Please try again\n")
                    continue
                else:
                    break
                
    driver = webdriver.Chrome('D:\\chdr.exe',options=chromeOptions,desired_capabilities=chromeOptions.to_capabilities())
    driver.implicitly_wait(10)
    driver.get('http://mygp.srv.gapac.com/c/cpgit/CPGITExpProj/bam/Master%20Team%20Documents/Forms/Data.aspx')
    element=driver.find_element_by_id("QCB1_Button2")
    element.click()
    driver.switch_to.frame(driver.find_element_by_xpath(".//iframe[@src='http://mygp.srv.gapac.com/c/cpgit/CPGITExpProj/bam/_layouts/15/Upload.aspx?List={88829755-EB61-4459-BBA3-6B3613F3AEB0}&RootFolder=&IsDlg=1']"));
    time.sleep(1)
    element1 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID,"ctl00_PlaceHolderMain_UploadDocumentSection_ctl05_InputFile")))
    #element1=driver.find_element_by_id("ctl00_PlaceHolderMain_UploadDocumentSection_ctl05_InputFile")
    element1.send_keys(file_path)
    time.sleep(3)
    driver.find_element_by_id("ctl00_PlaceHolderMain_ctl04_RptControls_btnOK").click()

    #time.sleep(4)
    ele_team = WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.XPATH, "//body/descendant::select[1]")))
    #ele_team = driver.find_element_by_xpath("//body/descendant::select[1]")
    ele_team.send_keys("Data")
    #time.sleep(1)
    ele_phase = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//body/descendant::select[3]")))
    ele_phase.send_keys(phase_path)
    #time.sleep(1.5)
    time.sleep(1.0)
    global ele_type
    ele_type = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//body/descendant::select[5]")))
    
    all_options = ele_type.find_elements_by_tag_name("option")
    #for option in all_options:
        #print(option.text)

    #return
    upath = None
                
    if choice == '1':
                    ele_type.send_keys('D2 - Load & Validation Summary - Material')
                    upath = 'D2 - Load & Validation Summary - Material'

    if choice == '2':
                    ele_type.send_keys('D2 - Load & Validation Summary - Vendor')
                    upath = 'D2 - Load & Validation Summary - Vendor'

    if choice == '3':
                    ele_type.send_keys('D2 - Load & Validation Summary - Customer')
                    upath = 'D2 - Load & Validation Summary - Customer'

    if choice == '4':
                    ele_type.send_keys('D2 - Load & Validation Summary - AC')
                    upath = 'D2 - Load & Validation Summary - AC'

    if choice == '5':
                    ele_type.send_keys('D2 - Load & Validation Summary - CTM')
                    upath = 'D2 - Load & Validation Summary - CTM'

    if choice == '6':
                    ele_type.send_keys('D2 - Load & Validation Summary - DAT')
                    upath = 'D2 - Load & Validation Summary - DAT'

    if choice == '7':
                    ele_type.send_keys('D2 - Load & Validation Summary - NTO')
                    upath = 'D2 - Load & Validation Summary - NTO'

    if choice == '8':
                    ele_type.send_keys('D2 - Load & Validation Summary - OTC')
                    upath = 'D2 - Load & Validation Summary - OTC'

    if choice == '9':
                    ele_type.send_keys('D2 - Load & Validation Summary - PTP')
                    upath = 'D2 - Load & Validation Summary - PTP'

    if choice == '10':
                    ele_type.send_keys('D2 - Load & Validation Summary - RTR')
                    upath = 'D2 - Load & Validation Summary - RTR'

    if choice == '11':
                    ele_type.send_keys('D2 - Load & Validation Summary - SCM')
                    upath = 'D2 - Load & Validation Summary - SCM'

    if choice == '12':
                    ele_type.send_keys('D2 - Load & Validation Summary - WM')
                    upath = 'D2 - Load & Validation Summary - WM'

    time.sleep(0.5)
    ele_save = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//form[@id='aspnetForm']/descendant::input[@value='Save']")))

    ele_save.click()
    #ele_save = driver.find_element_by_xpath("//form[@id='aspnetForm']/descendant::input[@value='Save']").click()
                
    print("\n..  File Successfully Uploaded to: " + Fore.BLACK + Back.GREEN + team_path + "-> " + phase_path + "-> " + upath + Style.RESET_ALL + "\n")
    try:
        driver.close()
    except:
        return
                
                    

def upload_recon(a):
    global driver1
    if a == 0:

        while True:
            etl_id = input('..  Please enter your ETL Task ID/RICEF ID OR Press 0 to go back.\n>>  ')
            if etl_id == '0':
                return
            #else:
                #break
            
            #try:
            #    etl_id = int(etl_id)
            #    break
            #except:
            #    print('\n..  Please enter a valid ETL Task ID\n')

        #try:
            driver1 = webdriver.Chrome('D:\\chdr.exe',options=chromeOptions1,desired_capabilities=chromeOptions1.to_capabilities())
            driver1.get('http://mygp.srv.gapac.com/c/cpgit/CPGITExpProj/bam/Lists/Cutover%20Reconciliation%20Summary')
            element=driver1.find_element_by_id("inplaceSearchDiv_WPQ2_lsinput")
            element.clear()
            element.send_keys(etl_id)
            element.send_keys(Keys.ENTER)
            time.sleep(0.5)
            element.clear()
        #except:
            #print("\n..  Error Occured! Please try again")
            #return
        

    if a != 0:
        #try:
            driver1 = webdriver.Chrome('D:\\chdr.exe',options=chromeOptions1,desired_capabilities=chromeOptions1.to_capabilities())
            driver1.get('http://mygp.srv.gapac.com/c/cpgit/CPGITExpProj/bam/Lists/Cutover%20Reconciliation%20Summary')
            element=driver1.find_element_by_id("inplaceSearchDiv_WPQ2_lsinput")
            element.clear()
            element.send_keys(a)
            element.send_keys(Keys.ENTER)
            time.sleep(0.5)
            element.clear()
            """except:
            pass
            #print("\n..  Error Occured in Uploading Online Recon! Please try again")
            return"""    

 
    
      
s = None   
while True:
    if s is None:
        print(Fore.BLUE + Back.WHITE + 'Validation Package Uploader [Version D2_Prod]' + Style.RESET_ALL)
    print('\n..  Enter 1 to Create and Upload Validation Zip file to Sharepoint')
    print('..  Enter 2 to Upload an already created Zip file to Sharepoint')
    print('..  Enter 3 to Update Online Recon Summary in Sharepoint')
    print('..  Enter C to clear the Screen')
    print('..  Enter 0 to exit\n')

    s=input(">>  ")

    if s == '1':
        upload_and_create()

    if s == '2':
        upload_only(0)

    if s == '3':
        upload_recon(0)

    #if s == '4':
        #text_to_excel()
        
    if s is not None and s.upper() == 'C':
        s = None
        a = os.system('cls')
        continue

    if s == '0':
        exit()
        break
    
    else:
        if s not in ('1','2','3','0','C','c'):
            print("..  Invalid Entry! Please Try again\n")
